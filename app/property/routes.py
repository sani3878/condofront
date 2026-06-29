from flask import render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from . import property_bp
from ..helpers import query_one, query_all, get_db

def get_setup_status(user):
    """Check what's been set up for this user."""
    status = {
        'has_property': False,
        'has_rooms': False,
        'has_staff': False,
        'property': None,
        'room_count': 0,
        'staff_count': 0,
    }

    if not user.property_id:
        return status

    status['has_property'] = True
    status['property'] = query_one("""
        SELECT * FROM tblproperty WHERE idno = %s
    """, [user.property_id])

    room = query_one("""
        SELECT COUNT(*) as cnt FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
    """, [user.property_id])
    status['room_count']  = room['cnt'] if room else 0
    status['has_rooms']   = status['room_count'] > 0

    staff = query_one("""
        SELECT COUNT(*) as cnt FROM tbluser
        WHERE property_id = %s AND is_active = TRUE
    """, [user.property_id])
    status['staff_count'] = staff['cnt'] if staff else 0
    status['has_staff']   = status['staff_count'] > 0

    return status


@property_bp.route('/dashboard')
@login_required
def dashboard():
    status = get_setup_status(current_user)
    return render_template('property/dashboard.html',
        active_page='settings', status=status)


@property_bp.route('/setup', methods=['GET', 'POST'])
@login_required
def setup():
    if request.method == 'POST':
        property_name = request.form.get('property_name', '').strip()
        address       = request.form.get('address', '').strip()
        property_code = request.form.get('property_code', '').strip()

        if not property_name:
            flash('กรุณากรอกชื่อโครงการ', 'danger')
            return redirect(url_for('property.setup'))

        db  = get_db()
        cur = db.cursor()

        try:
            cur.execute("""
                INSERT INTO tblproperty
                    (customer_id, property_name, property_code,
                     address, is_active)
                VALUES (%s, %s, %s, %s, TRUE)
                RETURNING idno
            """, [current_user.customer_id,
                  property_name,
                  property_code or None,
                  address or None])
            property_id = cur.fetchone()['idno']

            # Link user to this property
            cur.execute("""
                UPDATE tbluser SET property_id = %s
                WHERE idno = %s
            """, [property_id, current_user.id])

            db.commit()

            # Update current user object
            current_user.property_id = property_id

            flash('สร้างโครงการสำเร็จแล้ว!', 'success')
            return redirect(url_for('property.dashboard'))

        except Exception as e:
            db.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return render_template('property/setup.html')


@property_bp.route('/rooms')
@login_required
def rooms():
    status = get_setup_status(current_user)
    room_list = query_all("""
        SELECT * FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
        ORDER BY building, room_no
    """, [current_user.property_id]) if current_user.property_id else []

    return render_template('property/rooms.html',
                           status=status,
                           rooms=room_list)


@property_bp.route('/users')
@login_required
def users():
    status = get_setup_status(current_user)
    staff = query_all("""
        SELECT u.*, r.role_name
        FROM tbluser u
        JOIN tblrole r ON u.role_id = r.idno
        WHERE u.property_id = %s AND u.is_active = TRUE
        ORDER BY r.idno, u.fullname
    """, [current_user.property_id]) if current_user.property_id else []

    return render_template('property/users.html',
        active_page='staff',
                           status=status,
                           staff=staff)


@property_bp.route('/rooms/add', methods=['POST'])
@login_required
def add_room():
    building = request.form.get('building', '').strip().upper()
    room_no  = request.form.get('room_no', '').strip()
    floor    = request.form.get('floor', '').strip()

    if not room_no:
        flash('กรุณากรอกหมายเลขห้อง', 'danger')
        return redirect(url_for('property.rooms'))

    # Check duplicate
    existing = query_one("""
        SELECT idno FROM tblroom
        WHERE property_id = %s
          AND room_no = %s
          AND (building = %s OR (building IS NULL AND %s = ''))
          AND is_active = TRUE
    """, [current_user.property_id, room_no,
          building or None, building])

    if existing:
        flash(f'ห้อง {building}{room_no} มีอยู่แล้วในระบบ', 'danger')
        return redirect(url_for('property.rooms'))

    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute("""
            INSERT INTO tblroom
                (property_id, building, room_no, floor, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
        """, [current_user.property_id,
              building or None,
              room_no,
              int(floor) if floor else None])
        db.commit()
        flash(f'เพิ่มห้อง {building}{room_no} สำเร็จแล้ว!', 'success')

    except Exception as e:
        db.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('property.rooms'))


@property_bp.route('/rooms/download-template')
@login_required
def download_template():
    import io
    import openpyxl
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rooms'

    # Headers
    ws['A1'] = 'building'
    ws['B1'] = 'room_no'
    ws['C1'] = 'floor'

    # Instructions row
    ws['A2'] = 'A'
    ws['B2'] = '101'
    ws['C2'] = '1'
    ws['A3'] = 'A'
    ws['B3'] = '102'
    ws['C3'] = '1'
    ws['A4'] = 'B'
    ws['B4'] = '101'
    ws['C4'] = '1'
    ws['A5'] = 'ก'
    ws['B5'] = '201'
    ws['C5'] = '2'

    # Style headers
    from openpyxl.styles import Font, PatternFill
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True, color='FFFFFF')
        ws[cell].fill = PatternFill('solid', fgColor='1E3A5F')

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='condofront_rooms_template.xlsx'
    )


@property_bp.route('/rooms/import', methods=['POST'])
@login_required
def import_rooms():
    import openpyxl

    file = request.files.get('file')
    if not file or file.filename == '':
        flash('กรุณาเลือกไฟล์ Excel', 'danger')
        return redirect(url_for('property.rooms'))

    try:
        wb      = openpyxl.load_workbook(file)
        ws      = wb.active
        rows    = list(ws.iter_rows(min_row=2, values_only=True))

        success = 0
        skipped = 0
        errors  = 0

        db  = get_db()
        cur = db.cursor()

        for row in rows:
            building = str(row[0]).strip().upper() if row[0] else ''
            room_no  = str(row[1]).strip()         if row[1] else ''
            floor    = row[2]

            if not room_no:
                skipped += 1
                continue

            # Check duplicate
            cur.execute("""
                SELECT idno FROM tblroom
                WHERE property_id = %s
                  AND room_no = %s
                  AND is_active = TRUE
            """, [current_user.property_id, room_no])

            if cur.fetchone():
                skipped += 1
                continue

            try:
                cur.execute("""
                    INSERT INTO tblroom
                        (property_id, building, room_no, floor, is_active)
                    VALUES (%s, %s, %s, %s, TRUE)
                """, [current_user.property_id,
                      building or None,
                      room_no,
                      int(floor) if floor else None])
                success += 1
            except Exception:
                errors += 1

        db.commit()

        msg = f'นำเข้าสำเร็จ {success} ห้อง'
        if skipped: msg += f' · ข้ามไป {skipped} ห้อง (ซ้ำหรือว่าง)'
        if errors:  msg += f' · ผิดพลาด {errors} ห้อง'

        flash(msg, 'success' if success > 0 else 'warning')

    except Exception as e:
        flash(f'ไม่สามารถอ่านไฟล์ได้: {str(e)}', 'danger')

    return redirect(url_for('property.rooms'))
