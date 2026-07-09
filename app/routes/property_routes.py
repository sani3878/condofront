from flask import render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from ..blueprints import property_bp
from ..helpers import query_one, query_all, get_db

def get_setup_status(user):
    """Check what's been set up for this user."""
    status = {
        'has_property': False,
        'has_rooms': False,
        'has_staff': False,
        'property': None,
        'room_count': 0,
        'staff_count': 0,
    }

    if not user.property_id:
        return status

    status['has_property'] = True
    status['property'] = query_one("""
        SELECT * FROM tblproperty WHERE idno = %s
    """, [user.property_id])

    room = query_one("""
        SELECT COUNT(*) as cnt FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
    """, [user.property_id])
    status['room_count']  = room['cnt'] if room else 0
    status['has_rooms']   = status['room_count'] > 0

    staff = query_one("""
        SELECT COUNT(*) as cnt FROM tbluser
        WHERE property_id = %s AND is_active = TRUE
    """, [user.property_id])
    status['staff_count'] = staff['cnt'] if staff else 0
    status['has_staff']   = status['staff_count'] > 0

    return status


@property_bp.route('/dashboard')
@login_required
def dashboard():
    # Reception/Security → redirect straight to their property detail
    if current_user.role_id in (2, 3) and current_user.property_id:
        return redirect(url_for('property.detail',
                                property_id=current_user.property_id))

    # Owner/Manager → show all properties under their customer account
    properties = query_all("""
        SELECT p.idno, p.property_name, p.property_code,
               p.address, p.is_active,
               (SELECT COUNT(*) FROM tblroom r
                WHERE r.property_id = p.idno
                AND r.is_active = TRUE) AS room_count,
               (SELECT COUNT(*) FROM tbluser u
                WHERE u.property_id = p.idno
                AND u.is_active = TRUE
                AND u.role_id IN (1,2,3)) AS staff_count
        FROM tblproperty p
        WHERE p.customer_id = %s
        ORDER BY p.idno
    """, [current_user.customer_id]) if current_user.customer_id else []

    # Package limits
    pkg = query_one("""
        SELECT pkg.package_name, pkg.max_room,
               COALESCE(pkg.max_property, 1) AS max_property
        FROM tblcustomer c
        LEFT JOIN tblpackage pkg ON c.package_id = pkg.idno
        WHERE c.idno = %s
    """, [current_user.customer_id]) if current_user.customer_id else None

    return render_template('property/dashboard.html',
        active_page='settings',
        properties=properties,
        pkg=pkg,
        prop_count=len(properties))


@property_bp.route('/detail/<int:property_id>')
@login_required
def detail(property_id):
    """Property detail — rooms + staff. All staff can see their own property."""

    # Security check — staff can only see their own property
    if current_user.role_id in (2, 3):
        if current_user.property_id != property_id:
            flash('ไม่มีสิทธิ์เข้าถึงโครงการนี้', 'danger')
            return redirect(url_for('property.detail',
                                    property_id=current_user.property_id))

    # Owner can see any property under their customer
    if current_user.role_id == 1 and current_user.customer_id:
        prop = query_one("""
            SELECT * FROM tblproperty
            WHERE idno = %s AND customer_id = %s
        """, [property_id, current_user.customer_id])
    else:
        prop = query_one("""
            SELECT * FROM tblproperty WHERE idno = %s
        """, [property_id])

    if not prop:
        flash('ไม่พบโครงการนี้', 'danger')
        return redirect(url_for('property.dashboard'))

    # Rooms
    rooms = query_all("""
        SELECT * FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
        ORDER BY building, room_no
    """, [property_id])

    # Staff
    staff = query_all("""
        SELECT u.idno, u.fullname, u.email, u.mobile,
               u.is_active, r.role_name
        FROM tbluser u
        JOIN tblrole r ON u.role_id = r.idno
        WHERE u.property_id = %s
        AND u.role_id IN (1,2,3)
        AND u.is_active = TRUE
        ORDER BY u.role_id, u.fullname
    """, [property_id])

    # Package limits for this property
    pkg = query_one("""
        SELECT pkg.max_room
        FROM tblsubscription s
        JOIN tblpackage pkg ON s.package_id = pkg.idno
        WHERE s.property_id = %s AND s.is_active = TRUE
        ORDER BY s.idno DESC LIMIT 1
    """, [property_id])

    max_room   = pkg['max_room'] if pkg else None
    room_count = len(rooms)
    can_add    = max_room is None or room_count < max_room

    return render_template('property/detail.html',
        active_page='settings',
        prop=prop,
        rooms=rooms,
        staff=staff,
        room_count=room_count,
        max_room=max_room,
        can_add=can_add)


@property_bp.route('/edit/<int:property_id>', methods=['GET', 'POST'])
@login_required
def edit_property(property_id):
    """Edit property name, code, address."""
    prop = query_one("""
        SELECT * FROM tblproperty
        WHERE idno = %s AND customer_id = %s
    """, [property_id, current_user.customer_id])

    if not prop:
        flash('ไม่พบโครงการนี้', 'danger')
        return redirect(url_for('property.dashboard'))

    if request.method == 'POST':
        name    = request.form.get('property_name', '').strip()
        code    = request.form.get('property_code', '').strip()
        address = request.form.get('address', '').strip()

        if not name:
            flash('กรุณากรอกชื่อโครงการ', 'danger')
            return redirect(request.url)

        db  = get_db()
        cur = db.cursor()
        cur.execute("""
            UPDATE tblproperty SET
                property_name = %s,
                property_code = %s,
                address = %s
            WHERE idno = %s
        """, [name, code or None, address or None, property_id])
        db.commit()
        flash('บันทึกข้อมูลโครงการสำเร็จ', 'success')
        return redirect(url_for('property.detail', property_id=property_id))

    return render_template('property/edit_property.html', prop=prop)



@login_required
def setup():
    existing_property = None
    if current_user.property_id:
        existing_property = query_one("""
            SELECT * FROM tblproperty WHERE idno = %s
        """, [current_user.property_id])

    if request.method == 'POST':
        property_name = request.form.get('property_name', '').strip()
        address       = request.form.get('address', '').strip()
        property_code = request.form.get('property_code', '').strip()

        if not property_name:
            flash('กรุณากรอกชื่อโครงการ', 'danger')
            return redirect(url_for('property.dashboard'))

        db  = get_db()
        cur = db.cursor()

        try:
            if existing_property:
                cur.execute("""
                    UPDATE tblproperty
                    SET property_name = %s,
                        property_code = %s,
                        address = %s
                    WHERE idno = %s
                """, [property_name, property_code or None,
                      address or None, existing_property['idno']])
                db.commit()
                flash('บันทึกข้อมูลโครงการสำเร็จแล้ว!', 'success')
                return redirect(url_for('property.dashboard'))

            # ── Check property limit before creating new ──────
            pkg = query_one("""
                SELECT pkg.max_property,
                       (SELECT COUNT(*) FROM tblproperty p
                        WHERE p.customer_id = %s
                        AND p.is_active = TRUE) AS current_count
                FROM tblcustomer c
                LEFT JOIN tblpackage pkg ON c.package_id = pkg.idno
                WHERE c.idno = %s
            """, [current_user.customer_id, current_user.customer_id])

            if pkg and pkg['max_property']:
                if pkg['current_count'] >= pkg['max_property']:
                    flash(f'แพ็กเกจของคุณรองรับสูงสุด {pkg["max_property"]} โครงการ '
                          f'กรุณาอัปเกรดแพ็กเกจ', 'danger')
                    return redirect(url_for('property.dashboard'))

            # Creating a new property
            cur.execute("""
                INSERT INTO tblproperty
                    (customer_id, property_name, property_code,
                     address, is_active)
                VALUES (%s, %s, %s, %s, TRUE)
                RETURNING idno
            """, [current_user.customer_id,
                  property_name,
                  property_code or None,
                  address or None])
            property_id = cur.fetchone()['idno']

            # Create subscription
            customer_pkg = query_one("""
                SELECT package_id FROM tblcustomer WHERE idno = %s
            """, [current_user.customer_id])

            package_id = customer_pkg['package_id'] if customer_pkg else None
            if not package_id:
                free_pkg = query_one("""
                    SELECT idno FROM tblpackage
                    WHERE package_name = 'Free' AND is_active = TRUE
                    LIMIT 1
                """)
                package_id = free_pkg['idno'] if free_pkg else None

            if package_id:
                cur.execute("""
                    INSERT INTO tblsubscription
                        (property_id, package_id, start_date, is_active)
                    VALUES (%s, %s, CURRENT_DATE, TRUE)
                """, [property_id, package_id])

            # Link user to this property
            cur.execute("""
                UPDATE tbluser SET property_id = %s
                WHERE idno = %s
            """, [property_id, current_user.id])

            db.commit()
            current_user.property_id = property_id

            flash('สร้างโครงการสำเร็จแล้ว!', 'success')
            return redirect(url_for('property.dashboard'))

        except Exception as e:
            db.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    # Get all properties under this customer + limits
    all_properties = query_all("""
        SELECT p.idno, p.property_name, p.property_code, p.is_active,
               (SELECT COUNT(*) FROM tblroom r
                WHERE r.property_id = p.idno AND r.is_active = TRUE) AS room_count,
               (SELECT COUNT(*) FROM tbluser u
                WHERE u.property_id = p.idno AND u.is_active = TRUE
                AND u.role_id IN (1,2,3)) AS staff_count
        FROM tblproperty p
        WHERE p.customer_id = %s
        ORDER BY p.idno
    """, [current_user.customer_id]) if current_user.customer_id else []

    pkg_limit = query_one("""
        SELECT pkg.max_property
        FROM tblcustomer c
        LEFT JOIN tblpackage pkg ON c.package_id = pkg.idno
        WHERE c.idno = %s
    """, [current_user.customer_id]) if current_user.customer_id else None

    max_property  = pkg_limit['max_property'] if pkg_limit else 1
    prop_count    = len(all_properties)
    can_add_more  = prop_count < max_property

    return render_template('property/setup.html',
        property=existing_property,
        all_properties=all_properties,
        prop_count=prop_count,
        max_property=max_property,
        can_add_more=can_add_more)


@property_bp.route('/rooms')
@login_required
def rooms():
    status = get_setup_status(current_user)
    room_list = query_all("""
        SELECT * FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
        ORDER BY building, room_no
    """, [current_user.property_id]) if current_user.property_id else []

    return render_template('property/rooms.html',
                           status=status,
                           rooms=room_list)


@property_bp.route('/users')
@login_required
def users():
    status = get_setup_status(current_user)
    staff = query_all("""
        SELECT u.*, r.role_name
        FROM tbluser u
        JOIN tblrole r ON u.role_id = r.idno
        WHERE u.property_id = %s AND u.is_active = TRUE
        ORDER BY r.idno, u.fullname
    """, [current_user.property_id]) if current_user.property_id else []

    # Subscription limit info
    max_user = None
    package_name = None
    if current_user.property_id:
        sub = query_one("""
            SELECT pkg.max_user, pkg.package_name
            FROM tblsubscription s
            JOIN tblpackage pkg ON s.package_id = pkg.idno
            WHERE s.property_id = %s AND s.is_active = TRUE
            ORDER BY s.idno DESC LIMIT 1
        """, [current_user.property_id])
        if sub:
            max_user = sub['max_user']
            package_name = sub['package_name']

    roles = query_all("SELECT * FROM tblrole ORDER BY idno")

    return render_template('property/users.html',
        active_page='staff',
        status=status,
        staff=staff,
        roles=roles,
        staff_count=len(staff),
        max_user=max_user,
        package_name=package_name)


@property_bp.route('/billing')
@login_required
def billing():
    status = get_setup_status(current_user)

    current_sub = None
    if current_user.property_id:
        current_sub = query_one("""
            SELECT s.idno AS sub_id, s.start_date, s.expire_date,
                   pkg.idno AS package_id, pkg.package_name, pkg.monthly_fee,
                   pkg.max_room, pkg.max_user, pkg.max_parcel,
                   COALESCE(pkg.max_property, 1) AS max_property
            FROM tblsubscription s
            JOIN tblpackage pkg ON s.package_id = pkg.idno
            WHERE s.property_id = %s AND s.is_active = TRUE
            ORDER BY s.idno DESC LIMIT 1
        """, [current_user.property_id])

    # Current usage against the plan's limits
    usage = {'room': 0, 'user': 0, 'parcel': 0, 'property': 0}
    if current_user.property_id:
        usage['room'] = query_one("""
            SELECT COUNT(*) AS cnt FROM tblroom
            WHERE property_id = %s AND is_active = TRUE
        """, [current_user.property_id])['cnt']

        usage['user'] = query_one("""
            SELECT COUNT(*) AS cnt FROM tbluser
            WHERE property_id = %s AND is_active = TRUE
        """, [current_user.property_id])['cnt']

        usage['parcel'] = query_one("""
            SELECT COUNT(*) AS cnt FROM tblparcel
            WHERE property_id = %s AND deleted_at IS NULL
        """, [current_user.property_id])['cnt']

        # Property count across entire customer account
        if current_user.customer_id:
            usage['property'] = query_one("""
                SELECT COUNT(*) AS cnt FROM tblproperty
                WHERE customer_id = %s AND is_active = TRUE
            """, [current_user.customer_id])['cnt']

    all_packages = query_all("""
        SELECT * FROM tblpackage
        WHERE is_active = TRUE
        ORDER BY monthly_fee
    """)

    return render_template('property/billing.html',
        active_page='settings',
        status=status,
        current_sub=current_sub,
        usage=usage,
        all_packages=all_packages)


@property_bp.route('/users/add', methods=['POST'])
@login_required
def add_user():
    from werkzeug.security import generate_password_hash
    import secrets

    # Get redirect target — property detail or users page
    redirect_property = request.form.get('redirect_property')

    if not current_user.property_id:
        flash('กรุณาตั้งค่าโครงการก่อน', 'danger')
        return redirect(url_for('property.dashboard'))

    fullname  = request.form.get('fullname', '').strip()
    email     = request.form.get('email', '').strip().lower()
    mobile    = request.form.get('mobile', '').strip()
    role_id   = request.form.get('role_id')
    password  = request.form.get('password', '').strip()

    # Use property_id from form if adding to a different property
    target_property = int(redirect_property) if redirect_property else current_user.property_id

    if not all([fullname, email, role_id]):
        flash('กรุณากรอกข้อมูลให้ครบ', 'danger')
        return redirect(url_for('property.detail', property_id=target_property))

    if query_one('SELECT idno FROM tbluser WHERE email = %s', [email]):
        flash('อีเมลนี้ถูกใช้งานแล้ว', 'danger')
        return redirect(url_for('property.detail', property_id=target_property))

    # Use provided password or generate temp
    temp_password = password if password else secrets.token_urlsafe(8)

    db  = get_db()
    cur = db.cursor()
    cur.execute("""
        INSERT INTO tbluser
            (customer_id, property_id, role_id,
             email, password_hash, fullname, mobile,
             is_active, email_verified)
        VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, TRUE)
    """, [current_user.customer_id, target_property, role_id,
          email, generate_password_hash(temp_password),
          fullname, mobile or None])
    db.commit()

    flash(f'เพิ่มพนักงานสำเร็จ! '
          f'รหัสผ่านชั่วคราว: {temp_password} '
          f'(กรุณาแจ้งพนักงานให้เปลี่ยนรหัสผ่านทันที)', 'success')
    return redirect(url_for('property.detail', property_id=target_property))


@property_bp.route('/users/deactivate/<int:user_id>', methods=['POST'])
@login_required
def deactivate_user(user_id):
    if user_id == current_user.id:
        flash('ไม่สามารถปิดใช้งานบัญชีตัวเองได้', 'danger')
        return redirect(url_for('property.users'))

    db  = get_db()
    cur = db.cursor()
    cur.execute("""
        UPDATE tbluser SET is_active = FALSE
        WHERE idno = %s AND property_id = %s
    """, [user_id, current_user.property_id])
    db.commit()

    flash('ปิดใช้งานพนักงานสำเร็จ', 'success')
    return redirect(url_for('property.users'))


@property_bp.route('/rooms/add', methods=['POST'])
@login_required
def add_room():
    building = request.form.get('building', '').strip().upper()
    room_no  = request.form.get('room_no', '').strip()
    floor    = request.form.get('floor', '').strip()

    if not room_no:
        flash('กรุณากรอกหมายเลขห้อง', 'danger')
        return redirect(url_for('property.rooms'))

    # Check duplicate
    existing = query_one("""
        SELECT idno FROM tblroom
        WHERE property_id = %s
          AND room_no = %s
          AND (building = %s OR (building IS NULL AND %s = ''))
          AND is_active = TRUE
    """, [current_user.property_id, room_no,
          building or None, building])

    if existing:
        flash(f'ห้อง {building}{room_no} มีอยู่แล้วในระบบ', 'danger')
        return redirect(url_for('property.rooms'))

    try:
        import secrets as _sec
        invite_code = _sec.token_urlsafe(6).upper()[:8]

        db  = get_db()
        cur = db.cursor()
        cur.execute("""
            INSERT INTO tblroom
                (property_id, building, room_no, floor, invite_code, is_active)
            VALUES (%s, %s, %s, %s, %s, TRUE)
        """, [current_user.property_id,
              building or None,
              room_no,
              int(floor) if floor else None,
              invite_code])
        db.commit()
        flash(f'เพิ่มห้อง {building}{room_no} สำเร็จแล้ว! รหัสเชิญ: {invite_code}', 'success')

    except Exception as e:
        db.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('property.rooms'))


@property_bp.route('/rooms/download-template')
@login_required
def download_template():
    import io
    import openpyxl
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rooms'

    # Headers
    ws['A1'] = 'building'
    ws['B1'] = 'room_no'
    ws['C1'] = 'floor'

    # Instructions row
    ws['A2'] = 'A'
    ws['B2'] = '101'
    ws['C2'] = '1'
    ws['A3'] = 'A'
    ws['B3'] = '102'
    ws['C3'] = '1'
    ws['A4'] = 'B'
    ws['B4'] = '101'
    ws['C4'] = '1'
    ws['A5'] = 'ก'
    ws['B5'] = '201'
    ws['C5'] = '2'

    # Style headers
    from openpyxl.styles import Font, PatternFill
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True, color='FFFFFF')
        ws[cell].fill = PatternFill('solid', fgColor='1E3A5F')

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='condofront_rooms_template.xlsx'
    )


@property_bp.route('/rooms/import', methods=['POST'])
@login_required
def import_rooms():
    import openpyxl

    file = request.files.get('file')
    if not file or file.filename == '':
        flash('กรุณาเลือกไฟล์ Excel', 'danger')
        return redirect(url_for('property.rooms'))

    try:
        wb      = openpyxl.load_workbook(file)
        ws      = wb.active
        rows    = list(ws.iter_rows(min_row=2, values_only=True))

        success = 0
        skipped = 0
        errors  = 0

        db  = get_db()
        cur = db.cursor()

        for row in rows:
            building = str(row[0]).strip().upper() if row[0] else ''
            room_no  = str(row[1]).strip()         if row[1] else ''
            floor    = row[2]

            if not room_no:
                skipped += 1
                continue

            # Check duplicate
            cur.execute("""
                SELECT idno FROM tblroom
                WHERE property_id = %s
                  AND room_no = %s
                  AND is_active = TRUE
            """, [current_user.property_id, room_no])

            if cur.fetchone():
                skipped += 1
                continue

            try:
                import secrets as _sec
                invite_code = _sec.token_urlsafe(6).upper()[:8]

                cur.execute("""
                    INSERT INTO tblroom
                        (property_id, building, room_no, floor, invite_code, is_active)
                    VALUES (%s, %s, %s, %s, %s, TRUE)
                """, [current_user.property_id,
                      building or None,
                      room_no,
                      int(floor) if floor else None,
                      invite_code])
                success += 1
            except Exception:
                errors += 1

        db.commit()

        msg = f'นำเข้าสำเร็จ {success} ห้อง'
        if skipped: msg += f' · ข้ามไป {skipped} ห้อง (ซ้ำหรือว่าง)'
        if errors:  msg += f' · ผิดพลาด {errors} ห้อง'

        flash(msg, 'success' if success > 0 else 'warning')

    except Exception as e:
        flash(f'ไม่สามารถอ่านไฟล์ได้: {str(e)}', 'danger')

    return redirect(url_for('property.rooms'))


@property_bp.route('/switch/<int:property_id>')
@login_required
def switch_property(property_id):
    """Switch active property for multi-property managers."""
    # Verify this property belongs to the same customer
    prop = query_one("""
        SELECT p.idno, p.property_name
        FROM tblproperty p
        WHERE p.idno = %s AND p.customer_id = %s AND p.is_active = TRUE
    """, [property_id, current_user.customer_id])

    if not prop:
        flash('ไม่พบโครงการนี้', 'danger')
        return redirect(url_for('main.home'))

    db  = get_db()
    cur = db.cursor()
    cur.execute("""
        UPDATE tbluser SET property_id = %s WHERE idno = %s
    """, [property_id, current_user.id])
    db.commit()

    # Update session
    current_user.property_id = property_id
    flash(f'เปลี่ยนโครงการเป็น {prop["property_name"]} แล้ว', 'success')
    return redirect(url_for('main.home'))


@property_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new_property():
    """Create an additional property under same customer."""
    # Check limit first
    pkg = query_one("""
        SELECT pkg.max_property,
               (SELECT COUNT(*) FROM tblproperty p
                WHERE p.customer_id = %s AND p.is_active = TRUE) AS current_count
        FROM tblcustomer c
        LEFT JOIN tblpackage pkg ON c.package_id = pkg.idno
        WHERE c.idno = %s
    """, [current_user.customer_id, current_user.customer_id])

    if pkg and pkg['max_property'] and pkg['current_count'] >= pkg['max_property']:
        flash(f'แพ็กเกจของคุณรองรับสูงสุด {pkg["max_property"]} โครงการ '
              f'กรุณาอัปเกรดแพ็กเกจ', 'danger')
        return redirect(url_for('property.dashboard'))

    if request.method == 'POST':
        property_name = request.form.get('property_name', '').strip()
        address       = request.form.get('address', '').strip()
        property_code = request.form.get('property_code', '').strip()

        if not property_name:
            flash('กรุณากรอกชื่อโครงการ', 'danger')
            return redirect(url_for('property.new_property'))

        db  = get_db()
        cur = db.cursor()

        try:
            cur.execute("""
                INSERT INTO tblproperty
                    (customer_id, property_name, property_code, address, is_active)
                VALUES (%s, %s, %s, %s, TRUE)
                RETURNING idno
            """, [current_user.customer_id, property_name,
                  property_code or None, address or None])
            property_id = cur.fetchone()['idno']

            # Copy subscription package from customer
            customer_pkg = query_one("""
                SELECT package_id FROM tblcustomer WHERE idno = %s
            """, [current_user.customer_id])

            if customer_pkg and customer_pkg['package_id']:
                cur.execute("""
                    INSERT INTO tblsubscription
                        (property_id, package_id, start_date, is_active)
                    VALUES (%s, %s, CURRENT_DATE, TRUE)
                """, [property_id, customer_pkg['package_id']])

            db.commit()
            flash(f'สร้างโครงการ {property_name} สำเร็จ! '
                  f'กรุณาเพิ่มพนักงานเพื่อจัดการโครงการนี้', 'success')
            return redirect(url_for('property.dashboard'))

        except Exception as e:
            db.rollback()
            flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return render_template('property/new_property.html',
        prop_count=pkg['current_count'] if pkg else 0,
        max_property=pkg['max_property'] if pkg else 1)


@property_bp.route('/room/add/<int:property_id>', methods=['POST'])
@login_required
def add_room_detail(property_id):
    """Add single room from property detail page."""
    building    = request.form.get('building', '').strip() or None
    room_no     = request.form.get('room_no', '').strip()
    owner_name  = request.form.get('owner_name', '').strip() or None
    owner_email = request.form.get('owner_email', '').strip() or None

    if not room_no:
        flash('กรุณากรอกเลขห้อง', 'danger')
        return redirect(url_for('property.detail', property_id=property_id))

    import secrets as _s
    invite_code = _s.token_urlsafe(6).upper()[:8]

    db  = get_db()
    cur = db.cursor()
    try:
        cur.execute("""
            INSERT INTO tblroom
                (property_id, building, room_no, owner_name,
                 owner_email, invite_code, is_active)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE)
        """, [property_id, building, room_no,
              owner_name, owner_email, invite_code])
        db.commit()
        flash(f'เพิ่มห้อง {building or ""}{room_no} สำเร็จ', 'success')
    except Exception as e:
        db.rollback()
        flash(f'เกิดข้อผิดพลาด: {str(e)}', 'danger')

    return redirect(url_for('property.detail', property_id=property_id))


@property_bp.route('/room/delete/<int:room_id>', methods=['POST'])
@login_required
def delete_room(room_id):
    """Soft delete a room."""
    db  = get_db()
    cur = db.cursor()
    room = query_one("SELECT property_id FROM tblroom WHERE idno = %s", [room_id])
    cur.execute("UPDATE tblroom SET is_active = FALSE WHERE idno = %s", [room_id])
    db.commit()
    flash('ลบห้องแล้ว', 'success')
    if room:
        return redirect(url_for('property.detail', property_id=room['property_id']))
    return redirect(url_for('property.dashboard'))


@property_bp.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    """Deactivate a staff user."""
    db  = get_db()
    cur = db.cursor()
    user = query_one("SELECT property_id FROM tbluser WHERE idno = %s", [user_id])
    cur.execute("""
        UPDATE tbluser SET is_active = FALSE
        WHERE idno = %s AND idno != %s
    """, [user_id, current_user.id])
    db.commit()
    flash('ลบพนักงานแล้ว', 'success')
    if user:
        return redirect(url_for('property.detail', property_id=user['property_id']))
    return redirect(url_for('property.dashboard'))


@property_bp.route('/room/reset-code/<int:room_id>', methods=['POST'])
@login_required
def reset_invite_code(room_id):
    """Reset invite code for a room — old code instantly invalid."""
    import secrets as _s
    new_code = _s.token_urlsafe(6).upper()[:8]

    db  = get_db()
    cur = db.cursor()
    room = query_one("""
        SELECT r.property_id FROM tblroom r WHERE r.idno = %s
    """, [room_id])

    if not room:
        flash('ไม่พบห้อง', 'danger')
        return redirect(url_for('property.dashboard'))

    cur.execute("""
        UPDATE tblroom SET
            invite_code     = %s,
            invite_used     = FALSE,
            invite_reset_at = NOW()
        WHERE idno = %s
    """, [new_code, room_id])
    db.commit()
    flash(f'รีเซ็ตรหัสเชิญสำเร็จ — รหัสใหม่: {new_code}', 'success')
    return redirect(url_for('property.detail', property_id=room['property_id']))


@property_bp.route('/rooms/print-labels')
@login_required
def print_room_labels():
    """Print room reference labels for all rooms."""
    rooms = query_all("""
        SELECT idno, building, room_no, invite_code, owner_name
        FROM tblroom
        WHERE property_id = %s AND is_active = TRUE
        ORDER BY building, room_no
    """, [current_user.property_id])

    property_info = query_one("""
        SELECT property_name FROM tblproperty WHERE idno = %s
    """, [current_user.property_id])

    return render_template('property/room_labels.html',
        rooms=rooms,
        property_name=property_info['property_name'] if property_info else '',
        hide_navbar=True)
